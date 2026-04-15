import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import json
import zipfile
from datetime import datetime
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os

# ── Page config 
st.set_page_config(
    page_title="Data Trend Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS 
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }

.stApp { background: #0d0d0d; color: #e8e8e8; }

.metric-card {
    background: #161616;
    border: 1px solid #2a2a2a;
    border-left: 3px solid #00ff88;
    padding: 1rem 1.25rem;
    border-radius: 2px;
    font-family: 'IBM Plex Mono', monospace;
}
.metric-label { font-size: 0.7rem; color: #888; letter-spacing: 0.1em; text-transform: uppercase; }
.metric-value { font-size: 1.6rem; font-weight: 600; color: #00ff88; margin-top: 0.2rem; }
.metric-sub   { font-size: 0.75rem; color: #555; margin-top: 0.1rem; }

.section-header {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    color: #555;
    border-bottom: 1px solid #222;
    padding-bottom: 0.4rem;
    margin-bottom: 1rem;
}

.outlier-badge {
    background: #1a0a0a;
    border: 1px solid #ff4444;
    color: #ff6666;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    padding: 0.2rem 0.5rem;
    border-radius: 2px;
    display: inline-block;
    margin: 0.2rem;
}

.export-section {
    background: #111;
    border: 1px solid #222;
    border-radius: 2px;
    padding: 1rem;
    margin-top: 0.5rem;
}

[data-testid="stSidebar"] { background: #0a0a0a; border-right: 1px solid #1e1e1e; }
[data-testid="stSidebar"] .stMarkdown { color: #999; }

div[data-testid="stSelectbox"] label,
div[data-testid="stMultiSelect"] label,
div[data-testid="stSlider"] label { color: #888 !important; font-size: 0.8rem !important; }

.stPlotlyChart { border: 1px solid #1e1e1e; border-radius: 2px; }
</style>
""", unsafe_allow_html=True)

# ── Helpers
PLOTLY_THEME = dict(
    paper_bgcolor="#0d0d0d",
    plot_bgcolor="#111111",
    font=dict(family="IBM Plex Mono", color="#aaa", size=11),
    xaxis=dict(gridcolor="#1e1e1e", linecolor="#2a2a2a", tickcolor="#333"),
    yaxis=dict(gridcolor="#1e1e1e", linecolor="#2a2a2a", tickcolor="#333"),
    colorway=["#00ff88", "#00aaff", "#ff6644", "#ffcc00", "#cc88ff"],
)

def apply_theme(fig):
    fig.update_layout(**PLOTLY_THEME)
    return fig

def detect_outliers_iqr(series: pd.Series):                            #Outlier Detection Algorithm using IQR
    q1, q3 = series.quantile(0.25), series.quantile(0.75)
    iqr = q3 - q1
    lower, upper = q1 - 1.5 * iqr, q3 + 1.5 * iqr
    mask = (series < lower) | (series > upper)
    return mask, lower, upper

def compute_rolling(df: pd.DataFrame, col: str, window: int):          #Computes Rolling Average
    return df[col].rolling(window=window, min_periods=1).mean()

def summary_stats(df: pd.DataFrame, numeric_cols):                     #Builds statistic table
    stats = df[numeric_cols].describe().T
    stats["skew"]     = df[numeric_cols].skew()
    stats["kurtosis"] = df[numeric_cols].kurtosis()
    stats["outlier_%"] = [
        round(detect_outliers_iqr(df[c])[0].mean() * 100, 2) for c in numeric_cols
    ]
    return stats

# ── Sample data generator, for when no CSV is in the program
def generate_sample_data():
    np.random.seed(42)
    n = 200
    dates = pd.date_range("2022-01-01", periods=n, freq="W")
    trend = np.linspace(100, 160, n)
    noise = np.random.normal(0, 5, n)
    seasonality = 10 * np.sin(np.linspace(0, 4 * np.pi, n))
    spikes = np.zeros(n)
    spikes[[30, 75, 140, 180]] = [40, -35, 50, -30]

    df = pd.DataFrame({
        "date":     dates,
        "revenue":  trend + noise + seasonality + spikes,
        "users":    np.random.randint(800, 2000, n) + np.linspace(0, 500, n).astype(int),
        "cost":     np.random.normal(60, 8, n) + np.linspace(0, 20, n),
        "category": np.random.choice(["Product A", "Product B", "Product C"], n),
        "region":   np.random.choice(["North", "South", "East", "West"], n),
    })
    df["revenue"] = df["revenue"].round(2)
    df["cost"]    = df["cost"].round(2)
    return df

# ── Export builders ───────────────────────────────────────────────────────────

def build_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode()

def build_excel(df: pd.DataFrame, stats_df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()

    # --- Sheet 1: Raw Data ---
    ws_data = wb.active
    ws_data.title = "Raw Data"
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(name="Calibri", bold=True, color="00FF88")
    border = Border(bottom=Side(style="thin", color="333333"))

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws_data.column_dimensions[get_column_letter(col_idx)].width = 16

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=val)

    # --- Sheet 2: Summary Stats ---
    ws_stats = wb.create_sheet("Summary Statistics")
    stats_reset = stats_df.reset_index()
    for col_idx, col_name in enumerate(stats_reset.columns, 1):
        cell = ws_stats.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = Font(bold=True, color="00FF88")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        ws_stats.column_dimensions[get_column_letter(col_idx)].width = 18

    for row_idx, row in enumerate(stats_reset.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=round(float(val), 4) if isinstance(val, (float, np.floating)) else val)
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="111111")

    # --- Sheet 3: Power BI Ready (flat, typed) ---
    ws_pbi = wb.create_sheet("PowerBI_Ready")
    pbi_df = df.copy()
    pbi_df["date"] = pbi_df["date"].dt.strftime("%Y-%m-%d")
    pbi_df["year"]  = pd.to_datetime(df["date"]).dt.year
    pbi_df["month"] = pd.to_datetime(df["date"]).dt.month
    pbi_df["week"]  = pd.to_datetime(df["date"]).dt.isocalendar().week.astype(int)

    for col_idx, col_name in enumerate(pbi_df.columns, 1):
        cell = ws_pbi.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0d47a1")
        ws_pbi.column_dimensions[get_column_letter(col_idx)].width = 16

    for row_idx, row in enumerate(pbi_df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws_pbi.cell(row=row_idx, column=col_idx, value=val)

    wb.save(buf)
    return buf.getvalue()

def build_pbit(df: pd.DataFrame) -> bytes:
    """
    Produces a .pbit-compatible ZIP containing:
      - DataModelSchema (minimal)
      - a CSV data snapshot
      - README with import instructions
    Power BI Desktop can load the CSV directly; the template defines
    the table/column structure so users just point it at the file.
    """
    buf = io.BytesIO()
    pbi_df = df.copy()
    pbi_df["date"]  = pbi_df["date"].dt.strftime("%Y-%m-%d")
    pbi_df["year"]  = pd.to_datetime(df["date"]).dt.year
    pbi_df["month"] = pd.to_datetime(df["date"]).dt.month
    pbi_df["week"]  = pd.to_datetime(df["date"]).dt.isocalendar().week.astype(int)

    schema = {
        "name": "DataTrendDashboard",
        "tables": [
            {
                "name": "DataTable",
                "columns": [{"name": c, "dataType": "string"} for c in pbi_df.columns],
            }
        ],
        "relationships": [],
        "cultures": [],
    }

    readme = """# Power BI Template — Data Trend Dashboard
    
## Import Instructions

### Option A — Use PowerBI_Ready.csv (Recommended)
1. Open Power BI Desktop
2. Home → Get Data → Text/CSV
3. Select `PowerBI_Ready.csv`
4. Click Transform Data, verify column types:
   - date  → Date
   - revenue, cost → Decimal Number
   - users → Whole Number
   - year, month, week → Whole Number
   - category, region → Text
5. Close & Apply
6. Build visuals!

### Option B — Use the .xlsx (PowerBI_Ready sheet)
1. Home → Get Data → Excel Workbook
2. Select the exported .xlsx
3. Choose the "PowerBI_Ready" sheet
4. Load / Transform as needed

## Suggested Visuals
- Line chart: date vs revenue (add rolling avg as second line)
- Bar chart: category vs revenue (grouped by region)
- Card visuals: total revenue, avg users, outlier count
- Scatter: revenue vs cost colored by category

## DAX Measures to Add
```dax
Total Revenue = SUM(DataTable[revenue])
Avg Users     = AVERAGE(DataTable[users])
Revenue YoY   = 
    VAR CY = [Total Revenue]
    VAR PY = CALCULATE([Total Revenue], SAMEPERIODLASTYEAR(DataTable[date]))
    RETURN DIVIDE(CY - PY, PY)
```
"""

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("DataModelSchema.json", json.dumps(schema, indent=2))
        z.writestr("PowerBI_Ready.csv", pbi_df.to_csv(index=False))
        z.writestr("README.txt", readme)

    return buf.getvalue()

def build_pdf(df: pd.DataFrame, stats_df: pd.DataFrame, fig_trend, fig_dist) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 fontSize=18, textColor=colors.HexColor("#00ff88"),
                                 spaceAfter=6)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#888888"))
    head_style  = ParagraphStyle("head", parent=styles["Heading2"],
                                 fontSize=11, textColor=colors.HexColor("#cccccc"),
                                 spaceBefore=12, spaceAfter=4)

    story = []
    story.append(Paragraph("Data Trend Analysis — Summary Report", title_style))
    story.append(Paragraph(f"Generated {datetime.now().strftime('%B %d, %Y  %H:%M')}", sub_style))
    story.append(Spacer(1, 0.2*inch))

    # Key metrics table
    numeric_cols = df.select_dtypes(include=np.number).columns.tolist()
    story.append(Paragraph("Summary Statistics", head_style))
    stats_reset = stats_df[["mean", "std", "min", "max", "outlier_%"]].reset_index()
    stats_reset.columns = ["Column", "Mean", "Std Dev", "Min", "Max", "Outlier %"]
    tbl_data = [stats_reset.columns.tolist()] + [
        [str(round(v, 3)) if isinstance(v, float) else str(v) for v in row]
        for row in stats_reset.itertuples(index=False)
    ]
    tbl = Table(tbl_data, hAlign="LEFT")
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0),  colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR",   (0,0), (-1,0),  colors.HexColor("#00ff88")),
        ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#111111"), colors.HexColor("#161616")]),
        ("TEXTCOLOR",   (0,1), (-1,-1), colors.HexColor("#cccccc")),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#2a2a2a")),
        ("PADDING",     (0,0), (-1,-1), 5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.2*inch))

    # Embed trend chart as image
    story.append(Paragraph("Trend Analysis", head_style))
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        fig_trend.write_image(tmp.name, width=900, height=320, scale=2)
        story.append(Image(tmp.name, width=9*inch, height=3*inch))

    story.append(Spacer(1, 0.15*inch))
    story.append(Paragraph("Distribution Overview", head_style))
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp2:
        fig_dist.write_image(tmp2.name, width=900, height=280, scale=2)
        story.append(Image(tmp2.name, width=9*inch, height=2.8*inch))

    doc.build(story)
    return buf.getvalue()

# ── App ───────────────────────────────────────────────────────────────────────
def main():
    # Sidebar
    with st.sidebar:
        st.markdown("### 📂 Data Source")
        upload = st.file_uploader("Upload CSV", type=["csv"])
        st.markdown("---")
        st.markdown("### ⚙️ Analysis")
        rolling_window = st.slider("Rolling Average Window", 2, 30, 7)
        st.markdown("---")
        st.markdown("<div style='color:#444;font-size:0.7rem;font-family:monospace'>DATA TREND DASHBOARD v1.0</div>", unsafe_allow_html=True)

    # Load data
    if upload:
        try:
            df = pd.read_csv(upload)
            date_cols = [c for c in df.columns if "date" in c.lower() or "time" in c.lower()]
            if date_cols:
                df[date_cols[0]] = pd.to_datetime(df[date_cols[0]], errors="coerce")
                df = df.rename(columns={date_cols[0]: "date"})
            st.success(f"✓ Loaded `{upload.name}` — {len(df):,} rows × {len(df.columns)} cols")
        except Exception as e:
            st.error(f"Parse error: {e}")
            df = generate_sample_data()
    else:
        df = generate_sample_data()
        st.info("No file uploaded — showing sample dataset (200 weeks of synthetic data)")

    numeric_cols = df.select_dtypes(include=np.number).columns.tolist()
    has_date     = "date" in df.columns

    # ── Header ────────────────────────────────────────────────────────────────
    st.markdown("""
    <div style='border-bottom:1px solid #1e1e1e;padding-bottom:0.75rem;margin-bottom:1.5rem'>
        <span style='font-family:IBM Plex Mono;font-size:1.5rem;font-weight:600;color:#e8e8e8'>
            DATA TREND ANALYSIS
        </span>
        <span style='font-family:IBM Plex Mono;font-size:0.75rem;color:#444;margin-left:1rem'>
            VISUALIZATION DASHBOARD
        </span>
    </div>
    """, unsafe_allow_html=True)

    # ── KPI row ───────────────────────────────────────────────────────────────
    stats = summary_stats(df, numeric_cols)
    kpi_cols = st.columns(len(numeric_cols))
    for i, col in enumerate(numeric_cols):
        outlier_mask, _, _ = detect_outliers_iqr(df[col])
        n_out = outlier_mask.sum()
        with kpi_cols[i]:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">{col}</div>
                <div class="metric-value">{df[col].mean():,.1f}</div>
                <div class="metric-sub">σ {df[col].std():,.1f} · {n_out} outliers</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Trend + Rolling Avg ───────────────────────────────────────────────────
    st.markdown('<div class="section-header">TREND ANALYSIS</div>', unsafe_allow_html=True)
    selected_col = st.selectbox("Select metric", numeric_cols, key="trend_col")

    if has_date:
        plot_df = df.sort_values("date").copy()
        plot_df["rolling"] = compute_rolling(plot_df, selected_col, rolling_window)
        outlier_mask, low, high = detect_outliers_iqr(plot_df[selected_col])
        plot_df["outlier"] = outlier_mask

        fig_trend = go.Figure()
        fig_trend.add_trace(go.Scatter(
            x=plot_df["date"], y=plot_df[selected_col],
            mode="lines", name=selected_col,
            line=dict(color="#00aaff", width=1, dash="dot"),
            opacity=0.5,
        ))
        fig_trend.add_trace(go.Scatter(
            x=plot_df["date"], y=plot_df["rolling"],
            mode="lines", name=f"{rolling_window}w Rolling Avg",
            line=dict(color="#00ff88", width=2),
        ))
        outliers = plot_df[plot_df["outlier"]]
        fig_trend.add_trace(go.Scatter(
            x=outliers["date"], y=outliers[selected_col],
            mode="markers", name="Outliers",
            marker=dict(color="#ff4444", size=7, symbol="x"),
        ))
        fig_trend.add_hrect(y0=high, y1=plot_df[selected_col].max()*1.05,
                            fillcolor="rgba(255,68,68,0.07)", line_width=0, annotation_text="upper IQR fence")
        fig_trend.add_hrect(y0=plot_df[selected_col].min()*0.95, y1=low,
                            fillcolor="rgba(255,68,68,0.07)", line_width=0, annotation_text="lower IQR fence")
        fig_trend.update_layout(height=360, **PLOTLY_THEME,
                                legend=dict(orientation="h", y=1.05))
        st.plotly_chart(fig_trend, use_container_width=True)
    else:
        st.warning("No date column detected — upload a CSV with a date/time column for trend charts.")
        fig_trend = go.Figure()

    # ── Distribution + Box ────────────────────────────────────────────────────
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown('<div class="section-header">DISTRIBUTION</div>', unsafe_allow_html=True)
        fig_dist = go.Figure()
        for c in numeric_cols:
            fig_dist.add_trace(go.Violin(y=df[c], name=c, box_visible=True,
                                         meanline_visible=True, opacity=0.8))
        fig_dist.update_layout(height=320, **PLOTLY_THEME)
        st.plotly_chart(fig_dist, use_container_width=True)

    with col_b:
        st.markdown('<div class="section-header">CORRELATION HEATMAP</div>', unsafe_allow_html=True)
        corr = df[numeric_cols].corr()
        fig_corr = go.Figure(go.Heatmap(
            z=corr.values, x=corr.columns, y=corr.index,
            colorscale=[[0,"#0d0d0d"],[0.5,"#004444"],[1,"#00ff88"]],
            text=corr.values.round(2), texttemplate="%{text}",
            showscale=True,
        ))
        fig_corr.update_layout(height=320, **PLOTLY_THEME)
        st.plotly_chart(fig_corr, use_container_width=True)

    # ── Category breakdown ────────────────────────────────────────────────────
    cat_cols = df.select_dtypes(include="object").columns.tolist()
    if cat_cols:
        st.markdown('<div class="section-header">CATEGORICAL BREAKDOWN</div>', unsafe_allow_html=True)
        c1, c2 = st.columns([1, 2])
        with c1:
            cat_col = st.selectbox("Group by", cat_cols)
            val_col = st.selectbox("Metric", numeric_cols, key="cat_val")
        with c2:
            grp = df.groupby(cat_col)[val_col].agg(["mean","sum","count"]).reset_index()
            fig_bar = px.bar(grp, x=cat_col, y="mean", color=cat_col,
                             color_discrete_sequence=["#00ff88","#00aaff","#ff6644","#ffcc00"])
            fig_bar.update_layout(height=260, showlegend=False, **PLOTLY_THEME)
            st.plotly_chart(fig_bar, use_container_width=True)

    # ── Summary Stats table ───────────────────────────────────────────────────
    st.markdown('<div class="section-header">SUMMARY STATISTICS</div>', unsafe_allow_html=True)
    st.dataframe(
        stats.style
            .format("{:.3f}")
            .background_gradient(subset=["outlier_%"], cmap="Reds", vmin=0, vmax=20)
            .background_gradient(subset=["mean"], cmap="Greens"),
        use_container_width=True,
    )

    # ── Outlier report ────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">OUTLIER REPORT (IQR METHOD)</div>', unsafe_allow_html=True)
    for col in numeric_cols:
        mask, lo, hi = detect_outliers_iqr(df[col])
        n = mask.sum()
        if n > 0:
            st.markdown(
                f'<span class="outlier-badge">⚠ {col}: {n} outliers '
                f'| fence [{lo:.2f}, {hi:.2f}]</span>',
                unsafe_allow_html=True,
            )

    # ── Exports ───────────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">EXPORT REPORTS</div>', unsafe_allow_html=True)

    with st.container():
        e1, e2, e3, e4 = st.columns(4)

        with e1:
            st.download_button(
                "⬇ Download CSV",
                data=build_csv(df),
                file_name=f"dashboard_export_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        with e2:
            excel_bytes = build_excel(df, stats)
            st.download_button(
                "⬇ Download Excel (.xlsx)",
                data=excel_bytes,
                file_name=f"dashboard_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with e3:
            pbit_bytes = build_pbit(df)
            st.download_button(
                "⬇ Power BI Package (.zip)",
                data=pbit_bytes,
                file_name=f"powerbi_package_{datetime.now().strftime('%Y%m%d')}.zip",
                mime="application/zip",
                use_container_width=True,
            )
            st.caption("Contains CSV + schema + DAX guide for Power BI Desktop")

        with e4:
            if has_date and selected_col in df.columns:
                try:
                    pdf_bytes = build_pdf(df, stats, fig_trend, fig_dist)
                    st.download_button(
                        "⬇ Download PDF Report",
                        data=pdf_bytes,
                        file_name=f"dashboard_report_{datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )
                except Exception as ex:
                    st.warning(f"PDF needs kaleido: `pip install kaleido` ({ex})")
            else:
                st.info("Upload CSV with date column to enable PDF export")

if __name__ == "__main__":
    main()
